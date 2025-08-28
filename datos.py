import tkinter as tk
from tkinter import messagebox, ttk
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# Rutas
DATA_FOLDER = "data"
EXCEL_FILE = os.path.join(DATA_FOLDER, "registro_viajes.xlsx")
REPORTS_FOLDER = "reportes"

# Inicialización de carpetas y archivo Excel
os.makedirs(DATA_FOLDER, exist_ok=True)
os.makedirs(REPORTS_FOLDER, exist_ok=True)

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws_trabajadores = wb.active
    ws_trabajadores.title = "Trabajadores"
    ws_trabajadores.append(["C.C.", "Nombres", "Apellidos", "Teléfono", "Estado"])
    wb.create_sheet("Inactivos").append(["C.C.", "Nombres", "Apellidos", "Teléfono"])
    wb.create_sheet("Viajes").append([
        "C.C.", "Nombre", "Fecha", "Placa", "Tonelaje", "ACPM (Gal)", "Precio ACPM", "Total ACPM", "Origen"
    ])
    wb.save(EXCEL_FILE)

# Función para cargar hojas

def cargar_hoja(nombre):
    wb = load_workbook(EXCEL_FILE)
    return wb[nombre], wb

# Función para guardar trabajador

def guardar_trabajador(cc, nombres, apellidos, telefono):
    ws, wb = cargar_hoja("Trabajadores")
    for row in ws.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == cc:
            messagebox.showerror("Error", "El trabajador ya existe.")
            return
    ws.append([cc, nombres, apellidos, telefono, "Activo"])
    wb.save(EXCEL_FILE)
    messagebox.showinfo("Éxito", "Trabajador guardado correctamente.")

# Función para obtener trabajadores activos

def obtener_trabajadores_activos():
    ws, _ = cargar_hoja("Trabajadores")
    return [f"{row[1]} {row[2]} ({row[0]})" for row in ws.iter_rows(min_row=2, values_only=True) if row[4] == "Activo"]

# Función para registrar viaje
lugares_usados = set()
placas_usadas = set()

def registrar_viaje(nombre_completo, fecha, placa, tonelaje, acpm, precio, origen):
    if not nombre_completo or "(" not in nombre_completo:
        messagebox.showerror("Error", "Selecciona un trabajador válido.")
        return
    cc = nombre_completo.split("(")[-1].replace(")", "").strip()
    nombre = nombre_completo.split("(")[0].strip()
    ws, wb = cargar_hoja("Viajes")
    total_acpm = round(float(acpm) * float(precio), 2) if acpm.lower() != "n/a" else "N/A"
    ws.append([cc, nombre, fecha, placa.upper(), tonelaje, acpm, precio, total_acpm, origen])
    wb.save(EXCEL_FILE)
    lugares_usados.add(origen)
    placas_usadas.add(placa.upper())
    messagebox.showinfo("Éxito", "Viaje registrado.")

# Función para ver viajes por trabajador

def ver_viajes():
    top = tk.Toplevel()
    top.title("Seleccionar Trabajador")
    top.geometry("400x250")
    top.lift()
    top.attributes('-topmost', True)
    top.after(0, lambda: top.attributes('-topmost', False))

    tk.Label(top, text="Selecciona un trabajador", font=("Helvetica", 14)).pack(pady=10)
    trabajadores = obtener_trabajadores_activos()
    combo = ttk.Combobox(top, values=trabajadores, font=("Helvetica", 14))
    combo.pack(pady=10)

    def mostrar_viajes():
        seleccionado = combo.get()
        if not seleccionado or "(" not in seleccionado:
            messagebox.showerror("Error", "Selecciona un trabajador válido.")
            return
        cc = seleccionado.split("(")[-1].replace(")", "").strip()

        ventana = tk.Toplevel()
        ventana.title("Historial de Viajes")
        ventana.geometry("1100x500")
        ventana.lift()
        ventana.attributes('-topmost', True)
        ventana.after(0, lambda: ventana.attributes('-topmost', False))

        columns = ("Nombre", "Fecha", "Placa", "Tonelaje", "ACPM", "Precio", "Total", "Origen")
        tree = ttk.Treeview(ventana, columns=columns, show="headings")
        for col in columns:
            ancho = 100 if col in ("Fecha", "Placa", "Tonelaje", "ACPM") else 120
            tree.heading(col, text=col)
            tree.column(col, width=ancho)

        total_ton = 0
        total_acpm = 0
        total_gasto = 0

        ws, _ = cargar_hoja("Viajes")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == cc:
                tree.insert("", "end", values=row[1:])
                try:
                    total_ton += float(row[4])
                    if str(row[5]).lower() != "n/a":
                        gal = float(row[5])
                        precio = float(row[6])
                        total_acpm += gal
                        total_gasto += gal * precio
                except Exception as e:
                    print("Error de conversión:", e)

        tree.pack(fill="both", expand=True)

        resumen = f"Total Toneladas: {total_ton:.2f} | Total ACPM (Gal): {total_acpm:.2f} | Gasto ACPM: ${total_gasto:,.0f}"
        tk.Label(ventana, text=resumen, font=("Helvetica", 12, "bold")).pack(pady=5)

        tk.Button(ventana, text="Salir", font=("Helvetica", 12), command=ventana.destroy).pack(pady=10)

    tk.Button(top, text="Ver Viajes", font=("Helvetica", 14), command=mostrar_viajes).pack(pady=10)
    tk.Button(top, text="Cancelar", font=("Helvetica", 14), command=top.destroy).pack(pady=5)

# Función para exportar reportes

def exportar_reporte(tipo):
    ws, _ = cargar_hoja("Viajes")
    registros = [row for row in ws.iter_rows(min_row=2, values_only=True)]
    headers = [cell.value for cell in cargar_hoja("Viajes")[0][1]]
    if tipo == "Mes":
        meses = {}
        for r in registros:
            mes = datetime.strptime(r[2], "%Y-%m-%d").strftime("%Y-%m")
            if mes not in meses:
                meses[mes] = []
            meses[mes].append(r)
        for mes, datos in meses.items():
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.append(headers)
            for d in datos:
                ws_out.append(d)
            wb_out.save(os.path.join(REPORTS_FOLDER, f"viajes_mes_{mes}.xlsx"))
    elif tipo == "Trabajador":
        trabajadores = {}
        for r in registros:
            cc = r[0]
            if cc not in trabajadores:
                trabajadores[cc] = []
            trabajadores[cc].append(r)
        for cc, datos in trabajadores.items():
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.append(headers)
            for d in datos:
                ws_out.append(d)
            nombre_archivo = datos[0][1].replace(" ", "_") + f"_{cc}.xlsx"
            wb_out.save(os.path.join(REPORTS_FOLDER, nombre_archivo))
    messagebox.showinfo("Listo", f"Reporte por {tipo.lower()} generado en la carpeta 'reportes'.")

# Interfaz principal

root = tk.Tk()
root.title("Administrador de Viajes")
root.geometry("600x500")
font_big = ("Helvetica", 14)

# Widgets
btn1 = tk.Button(root, text="Agregar Trabajador", font=font_big, command=lambda: ventana_trabajador())
btn2 = tk.Button(root, text="Registrar Viaje", font=font_big, command=lambda: ventana_viaje())
btn3 = tk.Button(root, text="Ver Viajes", font=font_big, command=ver_viajes)
btn4 = tk.Button(root, text="Exportar Reportes por Mes", font=font_big, command=lambda: exportar_reporte("Mes"))
btn5 = tk.Button(root, text="Exportar Reportes por Trabajador", font=font_big, command=lambda: exportar_reporte("Trabajador"))
btn6 = tk.Button(root, text="Salir", font=font_big, command=root.destroy)

btn1.pack(pady=10)
btn2.pack(pady=10)
btn3.pack(pady=10)
btn4.pack(pady=10)
btn5.pack(pady=10)
btn6.pack(pady=10)

# Ventanas secundarias

def ventana_trabajador():
    top = tk.Toplevel()
    top.title("Agregar Trabajador")
    top.geometry("400x300")
    top.lift()
    top.attributes('-topmost', True)
    top.after(0, lambda: top.attributes('-topmost', False))
    tk.Label(top, text="C.C.", font=font_big).pack()
    ent_cc = tk.Entry(top, font=font_big)
    ent_cc.pack()
    tk.Label(top, text="Nombres", font=font_big).pack()
    ent_nom = tk.Entry(top, font=font_big)
    ent_nom.pack()
    tk.Label(top, text="Apellidos", font=font_big).pack()
    ent_ape = tk.Entry(top, font=font_big)
    ent_ape.pack()
    tk.Label(top, text="Teléfono", font=font_big).pack()
    ent_tel = tk.Entry(top, font=font_big)
    ent_tel.pack()

    def guardar():
        guardar_trabajador(ent_cc.get(), ent_nom.get(), ent_ape.get(), ent_tel.get())

    tk.Button(top, text="Guardar", font=font_big, command=guardar).pack(pady=10)
    tk.Button(top, text="Salir", font=font_big, command=top.destroy).pack()

def ventana_viaje():
    top = tk.Toplevel()
    top.title("Registrar Viaje")
    top.geometry("500x700")
    top.lift()
    top.attributes('-topmost', True)
    top.after(0, lambda: top.attributes('-topmost', False))

    trabajadores = obtener_trabajadores_activos()

    tk.Label(top, text="Trabajador", font=font_big).pack()
    combo = ttk.Combobox(top, values=trabajadores, font=font_big)
    combo.pack()

    tk.Label(top, text="Fecha (YYYY-MM-DD)", font=font_big).pack()
    ent_fecha = tk.Entry(top, font=font_big)
    ent_fecha.insert(0, datetime.now().strftime("%Y-%m-%d"))
    ent_fecha.pack()

    tk.Label(top, text="Placa", font=font_big).pack()
    ent_placa = ttk.Combobox(top, values=list(placas_usadas), font=font_big)
    ent_placa.pack()

    tk.Label(top, text="Tonelaje", font=font_big).pack()
    ent_ton = tk.Entry(top, font=font_big)
    ent_ton.pack()

    var_tanqueo = tk.IntVar()
    check_acpm = tk.Checkbutton(top, text="¿Tanqueó?", variable=var_tanqueo, font=font_big)
    check_acpm.pack()

    tk.Label(top, text="ACPM (Gal)", font=font_big).pack()
    ent_acpm = tk.Entry(top, font=font_big, state="disabled")
    ent_acpm.pack()

    tk.Label(top, text="Precio ACPM por galón", font=font_big).pack()
    ent_precio = tk.Entry(top, font=font_big, state="disabled")
    ent_precio.pack()

    def toggle_acpm():
        if var_tanqueo.get():
            ent_acpm.config(state="normal")
            ent_precio.config(state="normal")
        else:
            ent_acpm.delete(0, tk.END)
            ent_precio.delete(0, tk.END)
            ent_acpm.insert(0, "N/A")
            ent_precio.insert(0, "N/A")
            ent_acpm.config(state="disabled")
            ent_precio.config(state="disabled")

    var_tanqueo.trace_add("write", lambda *args: toggle_acpm())
    toggle_acpm()

    tk.Label(top, text="Origen del Viaje", font=font_big).pack()
    ent_origen = ttk.Combobox(top, values=list(lugares_usados), font=font_big)
    ent_origen.pack()

    def guardar():
        registrar_viaje(
            combo.get(), ent_fecha.get(), ent_placa.get(), ent_ton.get(),
            ent_acpm.get(), ent_precio.get(), ent_origen.get()
        )

    tk.Button(top, text="Guardar Viaje", font=font_big, command=guardar).pack(pady=20)
    tk.Button(top, text="Salir", font=font_big, command=top.destroy).pack()

root.mainloop()
